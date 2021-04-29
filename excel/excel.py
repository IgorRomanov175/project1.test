import openpyxl


book = openpyxl.Workbook()

sheet = book.active
sheet.title = "test_title"
d = sheet.cell(row=4, column=2, value="hello")

sheet['A2'] = 100

sheet.merge_cells('A2:A4')

book.save('my_book.xlsx')
book.close()

